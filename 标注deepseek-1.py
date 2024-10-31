import openpyxl
from openai import OpenAI

# 初始化 OpenAI 客户端
client = OpenAI(api_key="", base_url="")

# 读取 Excel 文件
input_file_path = r""
output_file_path = r""

# 打开 Excel 文件
workbook = openpyxl.load_workbook(input_file_path)
sheet = workbook.active

# 遍历每一行，从第二行开始
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
    fact_description = row[1]  # 第二列的内容
    
    # 构建请求数据
    prompt = f"请帮我完成如下工作，我将给你一段事实描述，请根据我的事实描述，打三个标签，分别是    '000'， '001', '010', '011', '100', '101', '110', '111'，你只能输出这八个的其中之一。接下来我将给你具体的事实描述：{fact_description}"

    # 发送请求到 API
    response = client.chat.completions.create(
        model="",
        messages=[
            {"role": "system", "content": "You are a helpful assistant"},
            {"role": "user", "content": prompt},
        ],
        stream=False
    )
    
    # 解析响应
    if response.choices:
        result = response.choices[0].message.content.strip()
        # 假设结果是 "101" 这样的字符串
        if len(result) == 3:
            sheet.cell(row=row_idx, column=3, value=int(result[0]))  # Rerouting
            sheet.cell(row=row_idx, column=4, value=int(result[1]))  # Sanction
            sheet.cell(row=row_idx, column=5, value=int(result[2]))  # Trade-war
    else:
        print(f"Error processing row {row_idx}: No response from API")

# 保存输出文件
workbook.save(output_file_path)
print("Processing complete. Output saved to:", output_file_path)